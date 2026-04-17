from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

import pandas as pd
import xlrd
from openpyxl import load_workbook


DIRECT_CONTROLLED_CITIES = {"北京市", "上海市", "天津市", "重庆市"}
REGION_DELIMITERS = r"[-/\\|,，、\s]+"
PROVINCE_PATTERN = re.compile(r"^(.*?(?:省|自治区|特别行政区|市))")
CITY_PATTERN = re.compile(r"^(.*?(?:自治州|地区|盟|市))")
DISTRICT_PATTERN = re.compile(r"^(.*?(?:区|县|旗|市))")


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null"} or text == "-":
        return None
    return text


def normalize_identifier(value: Any) -> str | None:
    if value is None:
        return None
    if pd.isna(value):
        return None

    if isinstance(value, bool):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if pd.isna(value):
            return None
        if value.is_integer():
            return str(int(value))
        return format(Decimal(str(value)).normalize(), "f").rstrip("0").rstrip(".")

    text = clean_text(value)
    if not text:
        return None
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\u3000", "").replace(" ", "")
    if not text:
        return None

    normalized = text
    if re.fullmatch(r"[+-]?\d+\.0+", normalized):
        normalized = normalized.split(".", 1)[0]
    else:
        try:
            decimal_value = Decimal(normalized)
            if decimal_value == decimal_value.to_integral():
                normalized = str(decimal_value.quantize(Decimal("1")))
            else:
                normalized = format(decimal_value.normalize(), "f").rstrip("0").rstrip(".")
        except (InvalidOperation, ValueError):
            pass

    normalized = normalized.lstrip("+")
    return normalized or None


def parse_datetime(value: Any) -> datetime | None:
    text = clean_text(value)
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None
    if isinstance(parsed, pd.Timestamp):
        return parsed.to_pydatetime()
    return parsed


def parse_date(value: Any) -> date | None:
    dt = parse_datetime(value)
    return dt.date() if dt else None


def parse_float(value: Any) -> float:
    text = clean_text(value)
    if not text:
        return 0.0
    normalized = re.sub(r"[^\d.\-]", "", text)
    if not normalized:
        return 0.0
    try:
        return float(normalized)
    except ValueError:
        return 0.0


def dump_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def normalize_header(value: Any) -> str:
    text = clean_text(value) or ""
    text = text.replace("（", "(").replace("）", ")").replace("：", ":")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def normalize_headers(headers: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for raw in headers:
        base = normalize_header(raw) or "column"
        count = seen.get(base, 0)
        seen[base] = count + 1
        result.append(base if count == 0 else f"{base}.{count}")
    return result


def safe_stage_name(path: Path, sha: str, suffix: str) -> str:
    stem = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "_", path.stem).strip("_") or "orders"
    return f"{stem}__{sha[:12]}.{suffix}"


def detect_excel_format(path: Path) -> str | None:
    try:
        return xlrd.inspect_format(path=path.as_posix())
    except Exception:
        return None


def detect_largest_sheet(path: Path) -> str | None:
    excel_format = detect_excel_format(path) or path.suffix.lower().lstrip(".")
    if excel_format == "xls":
        workbook = xlrd.open_workbook(path.as_posix(), on_demand=True)
        largest = None
        largest_score = (-1, -1)
        for sheet in workbook.sheets():
            score = (sheet.nrows, sheet.ncols)
            if score > largest_score:
                largest = sheet.name
                largest_score = score
        workbook.release_resources()
        return largest

    if excel_format == "xlsx":
        with path.open("rb") as handle:
            workbook = load_workbook(handle, read_only=True, data_only=True)
            largest = None
            largest_score = (-1, -1)
            for sheet in workbook.worksheets:
                score = (sheet.max_row or 0, sheet.max_column or 0)
                if score > largest_score:
                    largest = sheet.title
                    largest_score = score
            workbook.close()
            return largest

    return None


def stage_excel_to_csv(path: Path, stage_path: Path) -> Path:
    stage_path.parent.mkdir(parents=True, exist_ok=True)
    target_sheet = detect_largest_sheet(path)
    if not target_sheet:
        raise ValueError(f"No worksheet found in {path}")

    excel_format = detect_excel_format(path) or path.suffix.lower().lstrip(".")
    if excel_format == "xls":
        workbook = xlrd.open_workbook(path.as_posix(), on_demand=True)
        sheet = workbook.sheet_by_name(target_sheet)
        with stage_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(normalize_headers(sheet.row_values(0)))
            for row_idx in range(1, sheet.nrows):
                values = sheet.row_values(row_idx)
                writer.writerow(["" if value is None else value for value in values])
        workbook.release_resources()
        return stage_path

    with path.open("rb") as handle:
        workbook = load_workbook(handle, read_only=True, data_only=True)
        sheet = workbook[target_sheet]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            workbook.close()
            raise ValueError(f"No header row found in {path}")
        with stage_path.open("w", encoding="utf-8-sig", newline="") as output:
            writer = csv.writer(output)
            writer.writerow(normalize_headers(list(header_row)))
            for values in rows:
                writer.writerow(["" if value is None else value for value in values])
        workbook.close()
        return stage_path


def stage_csv_to_csv(path: Path, stage_path: Path) -> Path:
    stage_path.parent.mkdir(parents=True, exist_ok=True)
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "latin1"):
        try:
            with path.open("r", encoding=encoding, newline="") as source:
                reader = csv.reader(source)
                header = next(reader, None)
                if header is None:
                    raise ValueError(f"CSV header missing: {path}")
                normalized_header = normalize_headers(list(header))
                with stage_path.open("w", encoding="utf-8-sig", newline="") as target:
                    writer = csv.writer(target)
                    writer.writerow(normalized_header)
                    for row in reader:
                        writer.writerow(row)
                return stage_path
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            continue
    if last_error:
        raise last_error
    raise ValueError(f"Failed to stage CSV: {path}")


def iter_csv_chunks(path: Path, chunk_size: int) -> Iterator[list[dict[str, Any]]]:
    for encoding in ("utf-8-sig", "utf-8", "gbk", "latin1"):
        try:
            for chunk in pd.read_csv(path, dtype=str, encoding=encoding, chunksize=chunk_size, keep_default_na=False):
                yield chunk.fillna("").to_dict(orient="records")
            return
        except UnicodeDecodeError:
            continue


def load_table(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows: list[dict[str, Any]] = []
        for chunk in iter_csv_chunks(path, chunk_size=50000):
            rows.extend(chunk)
        return rows

    if suffix in {".xlsx", ".xls"}:
        workbook = pd.read_excel(path, sheet_name=None, dtype=str)
        if not workbook:
            return []
        frame = max(workbook.values(), key=lambda item: (len(item.index), len(item.columns)))
        frame = frame.fillna("")
        frame.columns = normalize_headers(list(frame.columns))
        return frame.to_dict(orient="records")
    raise ValueError(f"Unsupported file type: {path.suffix}")


def first_region_line(value: str | None) -> str | None:
    if not value:
        return None
    for line in str(value).splitlines():
        text = clean_text(line)
        if text:
            return text
    return clean_text(value)


def _split_region_tokens(text: str) -> list[str]:
    return [part.strip() for part in re.split(REGION_DELIMITERS, text) if part.strip()]


def _extract_region_by_suffix(text: str) -> tuple[str | None, str | None, str | None]:
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return None, None, None

    province_match = PROVINCE_PATTERN.match(compact)
    if not province_match:
        return compact, None, None

    province = province_match.group(1)
    rest = compact[len(province):]
    if province in DIRECT_CONTROLLED_CITIES:
        district_match = DISTRICT_PATTERN.match(rest)
        district = district_match.group(1) if district_match else (rest or None)
        return province, province, district

    city_match = CITY_PATTERN.match(rest)
    if not city_match:
        return province, None, rest or None

    city = city_match.group(1)
    rest = rest[len(city):]
    district_match = DISTRICT_PATTERN.match(rest)
    district = district_match.group(1) if district_match else (rest or None)
    return province, city, district


def parse_region(value: str | None) -> tuple[str | None, str | None, str | None]:
    text = first_region_line(value)
    if not text:
        return None, None, None

    parts = _split_region_tokens(text)
    if len(parts) <= 1:
        return _extract_region_by_suffix(text)

    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    if len(parts) == 2:
        if parts[0] in DIRECT_CONTROLLED_CITIES:
            return parts[0], parts[0], parts[1]
        return parts[0], parts[1], None
    return parts[0], None, None


def days_between(left: date | None, right: date | None) -> int | None:
    if not left or not right:
        return None
    return (left - right).days


def safe_ratio(numerator: float, denominator: float) -> float:
    if not denominator:
        return 0.0
    return round(numerator / denominator, 4)


def infer_order_month_from_filename(file_name: str) -> str | None:
    text = clean_text(file_name) or ""
    match = re.search(r"(20\d{2})[^\d]?(\d{1,2})月", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    match = re.search(r"(20\d{2})[^\d]?(\d{1,2})", text)
    if match:
        month = int(match.group(2))
        if 1 <= month <= 12:
            return f"{match.group(1)}-{month:02d}"
    return None


def infer_order_month_from_value(value: Any) -> str | None:
    parsed = parse_datetime(value)
    if not parsed:
        return None
    return parsed.strftime("%Y-%m")
